# -*- coding: utf-8 -*-
from scrapy import Spider, Request
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from scrapy.http import TextResponse
import csv, xlsxwriter, xlrd, xlwt, time, calendar
from xlutils.copy import copy
import openpyxl
def RepresentsInt(s):
    try:
        int(s)
        return True
    except ValueError:
        return False

class corabastos_com_coSpider(Spider):
    name = "corabastos_com_co"
    start_url = 'https://www.corabastos.com.co/aNuevo/index.php?option=com_content&view=article&id=471&Itemid=290'
    domain1 = 'https://www.corabastos.com.co'

    driver = None
    conn = None
    total_message_count = 0

    ids = []
    all_date_data = []
    all_fields = []

    start_date = ''
    end_date = ''
    file_path = ''

    workbook_to_write = None
    sheet_to_write = None

    wb = None


    now_row_count = 1

    def start_requests(self):
        yield Request(self.start_url)

    def parse(self, response):

        # email = 'alvaromollica@gmail.com'
        # password = '26d0vh1u'
        # base_url = 'http://www.inmet.gov.br/projetos/rede/pesquisa/gera_serie_txt.php?&mRelEstacao={}&btnProcesso=serie&mRelDtInicio={}&mRelDtFim={}&mAtributos=,,,,,,,,,,1,,,,,'

        chrome_options = Options()
        chrome_options.add_argument("window-size=1500,1500")

        self.driver = webdriver.Chrome("chromedriver.exe", options=chrome_options)
        self.driver.set_page_load_timeout(300)

        # login #####################################
        self.driver.get(self.start_url)

        iframe = self.driver.find_element_by_xpath('//div[@itemprop="articleBody"]/iframe')
        self.driver.switch_to.frame(iframe)
        self.driver.find_element_by_xpath('//a[text()="Precio Promedio Por Grupo y Fecha"]').click()

        try:
            wait = WebDriverWait(self.driver, 30)
            wait.until(EC.visibility_of_element_located((By.XPATH, '//input[@name="datepicker"]')))
        except:
            pass

        self.driver.find_element_by_xpath('//select[@id="slcalidad"]/option[@value="02"]').click()

        self.driver.find_element_by_xpath('//select[@id="slgrupo3"]/option[@value="1"]').click()
        self.driver.find_element_by_xpath('//select[@id="slgrupo4"]/option[@value="8"]').click()
        time.sleep(1)

        self.driver.find_element_by_xpath('//select[@id="slarticulo3"]/option[@value="100100"]').click()
        self.driver.find_element_by_xpath('//select[@id="slarticulo4"]/option[@value="808008"]').click()

        self.driver.execute_script('document.getElementsByClassName("hasDatepicker")[0].removeAttribute("readonly")')
        self.driver.execute_script('document.getElementsByClassName("hasDatepicker")[1].removeAttribute("readonly")')

        self.wb = openpyxl.load_workbook(self.file_path)

        w_sheet = self.wb.active

        years = [2019, 2018, 2017, 2016, 2015, 2014, 2013, 2012, 2011, 2010]
        months = ['12', '11', '10', '09', '08', '07', '06', '05', '04', '03', '02', '01']
        for i, y in enumerate(years):
            for j, m in enumerate(months):
                if y == 2019:
                    if j < 2:
                        continue
                start_date = '{}-{}-01'.format(str(y), m)
                end_date = '{}-{}-{}'.format(str(y), m, str(calendar.monthrange(y, int(m))[1]))

                input_tag = self.driver.find_element_by_xpath('//input[@name="datepicker"]')
                input_tag.clear()
                input_tag.send_keys(start_date)

                input_tag = self.driver.find_element_by_xpath('//input[@name="datepicker2"]')
                input_tag.clear()
                input_tag.send_keys(end_date)



                self.driver.find_element_by_xpath('//input[@value="Consultar"]').click()

                try:
                    wait = WebDriverWait(self.driver, 30)
                    wait.until(EC.visibility_of_element_located((By.XPATH, '//img[@src="img/indicator_remembermilk_orange.gif"]')))
                except:
                    pass

                resp1 = TextResponse(url=self.driver.current_url,
                                            body=self.driver.page_source,
                                            encoding='utf-8')

                while True:
                    resp1 = TextResponse(url=self.driver.current_url,
                                            body=self.driver.page_source,
                                            encoding='utf-8')
                    if not resp1.xpath('//img[@src="img/indicator_remembermilk_orange.gif"]'):
                        # time.sleep(1)
                        break
                    time.sleep(1)

                tr_list = resp1.xpath('//div[@id="divReportes"]/table/tbody/tr[@onclick]')

                count = 1

                start_date_val = '{}-{}'.format(str(y), m)
                row_val = self.all_date_data.index(start_date_val) + 2

                col_of_field = 1
                for tr in tr_list:
                    product_name = tr.xpath('./td[1]/text()').extract_first()
                    price = tr.xpath('./td[2]/text()').re(r'[\d.,]+')[0]


                    try:
                        col_of_field = self.all_fields.index(product_name) + 2
                    except:
                        col_of_field = len(self.all_fields) + 2
                        self.all_fields.append(product_name)
                        w_sheet.cell(row=1, column=col_of_field).value = product_name


                    w_sheet.cell(row=row_val, column=col_of_field).value = str(price)
            #     if j == 1:
            #         break
            # break

        self.wb.save(self.file_path)







        # # login_url = self.driver.find_element_by_xpath('//div[@id="fale_conosco_main"]/div/iframe').get_attribute('src')
        # # self.driver.get(login_url)
        # #
        # # market = self.driver.find_element_by_xpath('//form[@id="login"]/input"]')
        # # actions = ActionChains(self.driver)
        # # actions.move_to_element(market).perform()
        #
        # self.driver.find_element_by_xpath('//input[@value=" Acessar "]').click()
        #
        # # book_to_read = xlrd.open_workbook(self.file_path)
        # # sheet_to_read = book_to_read.sheet_by_index(0)
        #
        #
        # # rb = xlrd.open_workbook(self.file_path)
        # # wb = copy(rb)
        # # w_sheet = wb.get_sheet(0)
        #
        # for m, id_val in enumerate(self.ids):
        #
        #     self.driver.get(base_url.format(id_val, self.start_date, self.end_date))
        #     data = self.driver.page_source.split('Estacao;Data;Hora;Precipitacao;')[-1].split('</pre>')[0].split('\n')
        #
        #     if 'Estacao;Data;Hora;Precipitacao;' not in self.driver.page_source:
        #         print('\n########################################')
        #         print("The data is not existing. Run again")
        #         print('########################################\n')
        #
        #         self.driver.close()
        #
        #         self.driver = webdriver.Chrome("chromedriver.exe", options=chrome_options)
        #         self.driver.set_page_load_timeout(300)
        #
        #         # login #####################################
        #         self.driver.get('http://www.inmet.gov.br/projetos/rede/pesquisa/inicio.php')
        #
        #         # login_url = self.driver.find_element_by_xpath('//div[@id="fale_conosco_main"]/div/iframe').get_attribute('src')
        #         # self.driver.get(login_url)
        #         #
        #         # market = self.driver.find_element_by_xpath('//form[@id="login"]/input"]')
        #         # actions = ActionChains(self.driver)
        #         # actions.move_to_element(market).perform()
        #
        #
        #         username_input = self.driver.find_element_by_xpath('//input[@type="text"]')
        #         username_input.send_keys(email)
        #
        #         password_input = self.driver.find_element_by_xpath('//input[@type="password"]')
        #         password_input.send_keys(password)
        #
        #         self.driver.find_element_by_xpath('//input[@value=" Acessar "]').click()
        #         self.driver.get(base_url.format(id_val, self.start_date, self.end_date))
        #         data = self.driver.page_source.split('Estacao;Data;Hora;Precipitacao;')[-1].split('</pre>')[0].split('\n')
        #         if 'Estacao;Data;Hora;Precipitacao;' not in self.driver.page_source:
        #             print('\n########################################')
        #             print("The data is not existing. Run again")
        #             print('########################################\n')
        #
        #         # continue
        #
        #     try:
        #         col_of_station = self.ids.index(id_val) + 2
        #     except:
        #         col_of_station = 2
        #
        #     statiiion_name = self.driver.page_source.split('Estação           :')[-1].split('\n')[0].strip()
        #     w_sheet.cell(row=1, column=col_of_station).value = statiiion_name
        #
        #     count_to_add = 1
        #     for n, data_row in enumerate(data):
        #         data_row = data_row.strip()
        #         if not data_row:
        #             continue
        #         temps = data_row.split(';')
        #         try:
        #             date_data = temps[1]
        #             rain_fall = temps[3]
        #         except:
        #             continue
        #
        #
        #         try:
        #             row_of_date = self.all_date_data.index(date_data) + 2
        #         except Exception as e:
        #             row_of_date = self.now_row_count + count_to_add
        #             # w_sheet.cell(row=row_of_date, column=1).value = str(date_data)
        #             self.all_date_data.append(str(date_data))
        #             # w_sheet.write(row_of_
        #         count_to_add += 1
        #
        #         # w_sheet.write(row_of_date, col_of_station, str(rain_fall))#write station id on first row
        #         # w_sheet.cell(row=row_of_date, column=col_of_station).value = 'dsfsdaf'
        #         w_sheet.cell(row=row_of_date, column=col_of_station).value = str(rain_fall)
        #     print('\n########################################')
        #     print(str(m + 1) + '. ' + statiiion_name + ' ------ Completed')
        #     print('########################################\n')
        #
        #     # if m == 10:
        #     #     break
        # self.wb.save(self.file_path)